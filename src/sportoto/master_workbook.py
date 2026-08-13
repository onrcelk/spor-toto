"""Import the user's Sport Toto Master workbook into training records."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from .real_training import HistoricalMatch, build_training_frame


@dataclass(frozen=True)
class MasterImportReport:
    workbook: str
    total_rows: int
    valid_rows: int
    skipped_rows: int
    periods: int
    competitions: tuple[str, ...]


def load_master_matches(path: Path | str) -> tuple[list[HistoricalMatch], MasterImportReport]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import SportToto Master.xlsx") from exc

    source = Path(path).expanduser()
    workbook = load_workbook(source, read_only=True, data_only=True)
    if "MASTER VERİ" not in workbook.sheetnames:
        raise ValueError("Workbook does not contain MASTER VERİ sheet")
    sheet = workbook["MASTER VERİ"]
    matches: list[HistoricalMatch] = []
    periods: set[str] = set()
    total = 0
    skipped = 0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        total += 1
        period, _number, date_text, home, away, score, result, *_ = row
        if not home or not away or not period:
            skipped += 1
            continue
        try:
            datetime.strptime(str(date_text), "%d.%m.%Y %H:%M")
            home_goals, away_goals = (int(x) for x in str(score).split("-", 1))
            if str(result) not in {"0", "1", "2"}:
                raise ValueError("invalid result")
        except (TypeError, ValueError):
            skipped += 1
            continue
        periods.add(str(period))
        matches.append(HistoricalMatch(
            date=str(date_text),
            home_team=str(home).strip(),
            away_team=str(away).strip(),
            home_goals=home_goals,
            away_goals=away_goals,
            source="SportToto Master.xlsx",
            competition="Sport Toto mixed",
        ))
    report = MasterImportReport(str(source), total, len(matches), skipped, len(periods), ("Sport Toto mixed",))
    return matches, report


def build_master_training_frame(path: Path | str):
    matches, report = load_master_matches(path)
    return build_training_frame(matches), report


__all__ = ["MasterImportReport", "load_master_matches", "build_master_training_frame"]
